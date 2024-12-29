from genericpath import isfile
import os
from openpyxl import Workbook, load_workbook, utils
from openpyxl.styles import Font
from typing import List, Dict, Tuple

from common_functions import make_nearest_values_tuple, flag_suspicious_files

def log_to_excel_and_flag_outliers(
    list_with_deltas: List[Tuple[str, str, str, str, int]], 
    excelfile: str, 
    main_folder_path: str, 
    flagged_folder_path: str, 
    flagged_dict: Dict[str, List[str]], 
    valid_dict: Dict[str, int]
) -> Dict[str, List[str]]:
    """
    Logs deltas to an Excel file and flags suspicious files based on various conditions.

    This function processes the given list of deltas and logs the relevant data into an 
    Excel file. It also flags suspicious files (such as duplicates or subject changes) 
    and updates a dictionary of flagged files accordingly.

    Parameters:
    - list_with_deltas: A list of tuples, where each tuple contains information about
      two files (file names, timestamps, and delta time between them).
    - excelfile: Path to the Excel file where data will be logged.
    - main_folder_path: Path to the main folder containing files to be processed.
    - flagged_folder_path: Path to the folder where flagged files should be moved.
    - flagged_dict: A dictionary that holds flagged file information.
    - valid_dict: A dictionary containing thresholds for flagging files. Keys include:
      "dups_less_than", "dups_or_sidechange_less_than", "range_more_than", "range_less_than", 
      and "subjchange_more_than".

    Returns:
    - flagged_dict: The updated dictionary containing flagged files.
    """

    # Create an Excel file and initialize worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Main_sheet"
    
    if isfile(excelfile):
        os.remove(excelfile)
    wb.save(excelfile)

    wbk = load_workbook(excelfile)
    wsheet = wbk["Main_sheet"]

    for d in list_with_deltas:
        for row in wsheet.iter_cols(min_row=1, min_col=1, max_row=1000, max_col=1):
            for n, cell in enumerate(row, start=1):
                if cell.value is not None:
                    continue
                else:
                    letter = utils.get_column_letter(cell.column)
                    # Log values into the worksheet
                    cell.value = d[0]
                    wsheet[f'B{n}'].value = d[1]
                    wsheet[f'C{n}'].value = d[2]
                    wsheet[f'D{n}'].value = d[3]
                    wsheet[f'E{n}'].value = d[4]

                    # Check if the delta meets various conditions and flag outliers
                    if d[-1] < valid_dict["dups_less_than"] * 60:
                        wsheet[f'I{n}'].value = "<1"  # Only duplicates allowed
                        if d[0][:8] != d[1][:8]:  # Checks if barcodes don't match
                            if d[0][:5] != d[1][:5]:  # Subject change check
                                wsheet[f'J{n}'].value = "<1 OUTLIER: Subj change instead of duplicate"
                                wsheet[f'J{n}'].font = Font(color="FFFF0000")
                                set_for_later_reference = make_nearest_values_tuple(d, list_with_deltas)
                                flagged_dict = flag_suspicious_files(flagged_folder_path, main_folder_path, d, 
                                                                     "1less_IsSubjChange_NotDup", set_for_later_reference, flagged_dict)
                            elif d[0][5:8] != d[1][5:8]:  # Side change check
                                wsheet[f'J{n}'].value = "<1 OK: Side change"
                                wsheet[f'J{n}'].font = Font(color="FF008000")
                        else:
                            wsheet[f'J{n}'].value = "<1 OK: DUPLICATE"
                            wsheet[f'J{n}'].font = Font(color="FF008000")

                    elif d[-1] < valid_dict["dups_or_sidechange_less_than"] * 60:
                        wsheet[f'H{n}'].value = "<2"  # Duplicates and sporadic side changes allowed
                        if d[0][:8] != d[1][:8]:
                            if d[0][:5] != d[1][:5]:  # Subject change check
                                wsheet[f'J{n}'].value = "<2 OUTLIER: Subj change instead of duplicates/sidechange"
                                wsheet[f'J{n}'].font = Font(color="FFFF0000")
                                set_for_later_reference = make_nearest_values_tuple(d, list_with_deltas)
                                flagged_dict = flag_suspicious_files(flagged_folder_path, main_folder_path, d, 
                                                                     "2less_IsSubjChange_NotDup", set_for_later_reference, flagged_dict)

                            elif d[0][5:8] != d[1][5:8]:
                                wsheet[f'J{n}'].value = "<2 OK: Side change"
                                wsheet[f'J{n}'].font = Font(color="FF0000FF")
                        else:
                            wsheet[f'J{n}'].value = "<2 OK: DUPLICATE"
                            wsheet[f'J{n}'].font = Font(color="FF008000")

                    elif d[-1] > valid_dict["range_more_than"] * 60 and d[-1] < valid_dict["range_less_than"] * 60:
                        wsheet[f'G{n}'].value = " >2<5 "
                        if d[0][:8] != d[1][:8]:
                            if d[0][:5] != d[1][:5]:
                                wsheet[f'J{n}'].value = ">2<5 OK Subj change"
                                wsheet[f'J{n}'].font = Font(color="FF0000FF")
                            elif d[0][5:8] != d[1][5:8]:
                                wsheet[f'J{n}'].value = ">2<5 OK: Side change"
                                wsheet[f'J{n}'].font = Font(color="FF008000")
                        else:
                            wsheet[f'J{n}'].value = ">2<5 OUTLIER: DUPLICATE instead of sidechange/subj change"
                            wsheet[f'J{n}'].font = Font(color="FFFF0000")
                            set_for_later_reference = make_nearest_values_tuple(d, list_with_deltas)
                            flagged_dict = flag_suspicious_files(flagged_folder_path, main_folder_path, d, 
                                                                 "2more5less_IsDupl_NotSubjSidechange", set_for_later_reference, flagged_dict)

                    elif d[-1] > valid_dict["subjchange_more_than"] * 60:
                        wsheet[f'F{n}'].value = ">3 "
                        if d[0][:8] != d[1][:8]:
                            if d[0][:5] != d[1][:5]:
                                wsheet[f'J{n}'].value = ">3 OK: Subj change"
                                wsheet[f'J{n}'].font = Font(color="FF008000")
                            elif d[0][5:8] != d[1][5:8]:
                                wsheet[f'J{n}'].value = ">3 OK: Side change "
                                wsheet[f'J{n}'].font = Font(color="FF0000FF")
                        else:
                            wsheet[f'J{n}'].value = ">3 OUTLIER: DUPLICATE instead of sub change/side change"
                            wsheet[f'J{n}'].font = Font(color="FFFF0000")
                            set_for_later_reference = make_nearest_values_tuple(d, list_with_deltas)
                            flagged_dict = flag_suspicious_files(flagged_folder_path, main_folder_path, d, 
                                                                 "3more_IsDup_NotSubjSideChange", set_for_later_reference, flagged_dict)
                    break
            break
        wbk.save(excelfile)
    
    return flagged_dict
