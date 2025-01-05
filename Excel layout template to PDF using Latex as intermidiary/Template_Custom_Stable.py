from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook
import Internal_Imports_Stable as Imp
from Common_Functions_Stable import create_Latex_document
from Latex_File_Create_Stable import create_final_latex_file

from typing import List, Dict, Iterator, Tuple

def customize(
    area_sorted: List[str],
    time_sorted: List[str],
    sub_sorted: List[str],
    rownames: List[str],
    colnames: List[str],
    Main_mapping_dict: Dict[Tuple[str, str], str],
    random_iter: Iterator[Tuple[str, str]],
    Imp: Imp
) -> str:
    """
    Customizes the layout of a study template based on user preferences and study requirements.

    Args:
        area_sorted (List[str]): Sorted list of area names.
        time_sorted (List[str]): Sorted list of time points.
        sub_sorted (List[str]): Sorted list of subject IDs.
        rownames (List[str]): List of row names for the template.
        colnames (List[str]): List of column names for the template.
        Main_mapping_dict (Dict[Tuple[str, str], str]): Dictionary mapping tuples of (subject, area) to a random code or value.
        random_iter (Iterator[Tuple[str, str]]): Iterator of tuples containing file paths and corresponding random codes.
        Imp (Imp): An instance of the `Internal_Imports_Stable` module containing necessary import details like file paths and settings.

    Returns:
        str: The path to the generated LaTeX file.
    """
    next_label_count = "".join(rownames).count("next")
    
    # Reads the modified template
    wb = load_workbook(Imp.excelfile)
    ws4 = wb['Modify_Template_here']

    nr_col = len(colnames)
    nr_row = len(rownames)

    max_height = round(1.0 / (1.15 * nr_row), 2)
    if next_label_count > 0:
        max_height = round((1.0 * next_label_count) / (1.25 * nr_row), 2)
    
    max_width = round(1.0 / (1.2 * nr_col), 2)
        
    # Create LaTeX document with customized layout
    lax_document, new_page_alert = create_Latex_document(
        colnames, rownames, sub_sorted, max_height, max_width,
        area_sorted, time_sorted, ws4, Main_mapping_dict, Type="Custom", RandomList=random_iter
    )

    # Generate final LaTeX file
    custom_tex_file = create_final_latex_file(
        Imp.studynumber, Imp.header, Imp.pagestyle, Imp.hypersetup,
        lax_document, new_page_alert, colnames, Imp.Test_type, draft=Imp.draft_flag
    )
    
    return custom_tex_file
