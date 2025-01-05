import os
import subprocess
from time import sleep
import sys
import re

from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook, utils
from ImageAnalysis import Util

from Internal_Imports_Stable import InternalImport
import Common_Functions_Stable as CFS 

from Excel_Create_refactoring import create_excel 
from Randomization_Template_Stable import randomize_std, randomize_custom_rand_alllightingcode
from Randomization_Template_Transposed_Stable import randomize_transp
from B_Template_Stable import do_the_B_transform
from Column_Name_Per_Row_Template import give_each_row_columnames

from Insert_Description_Stable import Insert_description_file
from Helper_modules.file_renamers.Visia_file_rename_before_PDF_creation import rename_visia_files
from Templates import Standard, Transpose, Custom

def main() -> None:
    """
    The main function orchestrates the entire process, from image conversion to the final PDF generation.
    It handles tasks such as:
    1. Image file conversion and validation
    2. Missing file handling and renaming
    3. Excel template creation and user input handling
    4. Data randomization based on template flags
    5. Insertion of description pages
    6. PDF generation and archiving of results

    This function requires specific templates and files to be available on the system to work correctly.
    """
    Imp = InternalImport()

    # Convert and validate image files
    Imp.validated_path, resized_images = CFS.convert_and_scale_to_standard_jpgs(Imp.file_extension, 
                                                                                Imp.path_for_validation)
    
    subprocess.Popen(f"d: && start {Imp.validated_path}", shell=True)
    input("Conversion complete. Please check the created images in the opened folder and then press enter")

    # Check for transformed images correctness
    checker = input("Are the images transformed correctly...(y/n)?")
    while True:
        if checker.lower() == "y":
            file_extension = "*.jpg"
            break
        elif checker.lower() == "n": 
            print("Something wrong with the images. Stopping program..")
            sys.exit(1)
        checker = input("Are the images transformed correctly...(y/n)?")

    # Process and validate image files
    filenamelist = Util.getAllFiles(Imp.validated_path, "*.jpg", depth=-1)

    # Check for Visia files and handle renaming if necessary
    Visia_flag = 0
    fn_regx = re.compile(rf'{Imp.filename_mask}')
    visia_rgx_mask = re.compile(r"([0-9]*)_([A-Za-z0-9 ]*)_([a-zA-Z \-]*)_([a-zA-Z0-9 \-]*)")
    first_img_filename = [fn for fn in filenamelist if fn.endswith(".jpg") or fn.endswith(".JPG")][0].split("\\")[-1]
    
    if not re.match(fn_regx, first_img_filename):
        if visia_rgx_mask.search(first_img_filename).group(0):
            Visia_flag = 1
            rename_visia_files(Imp.validated_path, Imp.studynumber) 
        else:
            print("Regex match not matching for visia or std mask; check filename again")
            sys.exit(1)

    # Handle missing files and log dummy files
    filenamelist = Util.getAllFiles(Imp.validated_path, "*.jpg", depth=-1)
    dummy_list = CFS.replace_missing_barcodes_with_dummy(filenamelist, Visia_flag) 
    
    if dummy_list is not None: 
        filenamelist = Util.getAllFiles(Imp.validated_path, "*.jpg", depth=-1)  # re-read to include dummy images
        with open(os.path.join(Imp.validated_path, Imp.dummy_log_file), "w") as log_f:
            for item in dummy_list:
                log_f.write(item + "\n")
    
    # Perform final data cleaning
    CFS.remove_redundant_jpgs(Imp.validated_path)
    CFS.remove_counter_from_filenames(Imp.validated_path)
    CFS.remove_bitmaps_from_tif_conversion(Imp.validated_path)
    
    # Create Excel layout template
    dict_code = CFS.create_code_dict_for_excel_table(Imp.validated_path)
    area_code_list, time_code_list, area_sorted, time_sorted, sub_sorted, \
    filepaths, filenamelist, Main_mapping_dict, transfer_list = create_excel(Imp.validated_path, 
                                                                            Imp.excelfile, 
                                                                            dict_code, 
                                                                            Imp.filename_mask)

    # Open generated Excel file for user modifications
    open_path = os.getcwd()
    proc_excel = subprocess.Popen(f"start {os.path.join(open_path, Imp.excelfile)}", shell=True)
    proc_excel.wait()
    input("folder open..Press enter")

    # Read modified Excel file after user input
    wb = load_workbook(Imp.excelfile)
    ws4 = wb['Modify_Template_here']
    cell_a1 = ws4.cell(row=1, column=1).value  # checking if a * or § is found in this cell
    print(cell_a1)

    # Condition handling column names for each row
    column_row_name_list = []
    for coln, col in enumerate(ws4.columns):
        for cell in col:
            if cell.value == "§":  # Flag to include name above each image
                print("found")
                special_col_index = utils.get_column_letter(cell.column)
                column_row_name_list = []
                for row in ws4[special_col_index]:
                    if row.value in {"None", None, ""}:
                        break
                    column_row_name_list.append(row.value)
                column_row_name_list = column_row_name_list[1:]
                print(column_row_name_list)
                break
        if coln == 20:
            break

    # Process based on template flag (* or §)
    if cell_a1 == "*":
        randomized_files_iterator = None
        if Imp.randomfilepath is not None and Imp.isVisia:
            try:
                randomized_files_iterator = randomize_custom_rand_alllightingcode(filepaths, 
                                                          Imp.randomfilepath, 
                                                          area_sorted, time_sorted, sub_sorted)
            except IndexError:
                print("Custom - file List index out of range; Check the random file again for duplicates or Missing values")
                exit(1)
        elif Imp.randomfilepath is not None and not Imp.isVisia:
            try: 
                randomized_files_iterator = randomize_std(filepaths, 
                                                          Imp.randomfilepath, 
                                                          area_sorted, time_sorted, sub_sorted)
            except IndexError:
                print("Custom - file List index out of range; Check the random file again for duplicates or Missing values")
                exit(1)
                
        rownames, colnames = CFS.get_row_and_columns(ws4, 3, 2)
        
        cst = Custom(rownames, colnames)
        tex_file = cst.get_final_tex_file(sub_sorted, area_sorted, time_sorted,
                                          rownames, colnames, Main_mapping_dict, randomized_files_iterator, Imp)
        
    else:
        rownames, colnames = CFS.get_row_and_columns(ws4, 2, 1)
        if time_sorted[0] in rownames[0]:
            if Imp.randomfilepath is not None:
                try: 
                    randomized_files_iterator, filepathrandlist = randomize_transp(filepaths, 
                                                                                   Imp.randomfilepath, 
                                                                                   area_sorted, time_sorted, sub_sorted)
                    path_sep_list = [path[0] for path in filepathrandlist]
                    name_sep_file_list = [str_path.split("\\")[-1] for str_path in path_sep_list]
                    
                    tps = Transpose(area_code_list, time_code_list, 
                                    area_sorted, time_sorted, 
                                    rownames, 
                                    transfer_list, Imp)
                    tex_file = tps.get_final_tex_file(sub_sorted, area_sorted, time_sorted,
                                                      rownames, colnames, 
                                                      filepaths, filenamelist, 
                                                      randomized_files_iterator, Imp)
                except IndexError:
                    print("List index out of range; Check the random file again for duplicates or Missing values")
                    exit(1)
            else:
                randomized_files_iterator = None
                tps = Transpose(area_code_list, time_code_list, 
                                area_sorted, time_sorted,
                                rownames,
                                transfer_list, Imp)
                
                tex_file = tps.get_final_tex_file(sub_sorted, area_sorted, time_sorted,
                                                    rownames, colnames,
                                                    filepaths, filenamelist, 
                                                    randomized_files_iterator, Imp)

        else:
            if Imp.randomfilepath is not None:
                try: 
                    randomized_files_iterator = randomize_std(filepaths, 
                                                              Imp.randomfilepath, 
                                                              area_sorted, time_sorted, sub_sorted)
                except IndexError:
                    print("List index out of range; Check the random file again for duplicates or Missing values")
                    exit(1)
            else:
                randomized_files_iterator = None
            std = Standard(area_code_list, time_code_list, rownames)
            tex_file = std.get_final_tex_file(sub_sorted, area_sorted, time_sorted,
                                              rownames, colnames, 
                                              filepaths, filenamelist, 
                                              randomized_files_iterator, Imp)
            
    # Additional features/study-specific layouts
    if len(column_row_name_list) > 0:
        tex_file = give_each_row_columnames(column_row_name_list, tex_file)

    appended_beiers_fpath = None
    if Imp.Beirsdorf_randomfilepath is not None:
        _, appended_beiers_fpath = do_the_B_transform(tex_file, Imp.Beirsdorf_randomfilepath)

    # Insert description page into tex file
    if appended_beiers_fpath is not None:
        final_tex_file = Insert_description_file(appended_beiers_fpath.split("\\")[-1])
    else:
        final_tex_file = Insert_description_file(tex_file)

    # Generate PDF from tex file
    if os.path.isfile(final_tex_file):
        try:
            proc1 = subprocess.Popen(f"pdflatex -interaction=nonstopmode -halt-on-error {final_tex_file}  && pause", shell=True)
            proc1.wait()
            print("All files generated...starting archiving and cleaning process")
            sleep(1)
        except:
            os.startfile(final_tex_file)
            input("Press enter to exit...")
    
    CFS.archive_data(Imp.validated_path, 
                     Imp.studynumber, 
                     Imp.Test_type, 
                     Imp.file_extension)

if __name__ == '__main__':
    main()
