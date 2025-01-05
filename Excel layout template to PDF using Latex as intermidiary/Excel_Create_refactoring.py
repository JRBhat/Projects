from openpyxl import Workbook
import os
import re
from ImageAnalysis import Util
import sys
from typing import List, Dict, Tuple

def create_excel(mypath: str, excelfile: str, dict_code: Dict[str, str], filename_mask: str) -> Tuple[List[str], List[str], List[str], List[str], List[str], List[str], List[str], Dict[Tuple[str, str], str], List[str]]:
    """
    Creates an Excel file with several sheets, where the user can adjust layout and input additional names.
    
    The function processes images from a given directory, extracts subject, area, and time information
    based on filenames, and populates an Excel workbook with the relevant data. It organizes the data 
    in different sheets with appropriate templates, file paths, and codes for easy user modification.

    Args:
        mypath (str): Path to the directory containing image files.
        excelfile (str): Path to the output Excel file.
        dict_code (Dict[str, str]): Dictionary mapping area and time codes to corresponding values.
        filename_mask (str): Regular expression pattern used to match parts of filenames.

    Returns:
        Tuple[List[str], List[str], List[str], List[str], List[str], List[str], List[str], Dict[Tuple[str, str], str], List[str]]:
            - area_code_list: List of area codes.
            - time_code_list: List of time codes.
            - area_sorted: Sorted list of unique area identifiers.
            - time_sorted: Sorted list of unique time identifiers.
            - sub_sorted: Sorted list of unique subject identifiers.
            - filenamelist: List of full image file paths.
            - filenamelist_return: List of image filenames without paths.
            - Main_mapping_dict: Dictionary mapping (file path, subject number) tuples to full image paths.
            - transfer_list: List of file paths to be used in the final template.
    """

    subj = re.compile(r'S[0-9]{3}')
    area = re.compile(r'F[0-9]{2}')
    time = re.compile(r'T[0-9]{2}')
    file_key_mask = re.compile(rf'{filename_mask}')

    # List of file paths
    filenamelist = Util.getAllFiles(mypath, "*.jpg", depth=-1)

    ######################  SHEET 1 : Raw data storage ############################################
    wb = Workbook()
    ws1 = wb['Sheet']
    ws1.title = 'Raw_Data'

    for row in filenamelist:
        ws1.append((row, os.path.split(row)[-1]))
    wb.save(excelfile)

    ######################  SHEET 2 : Stores full template  #######################################
    ws2 = wb.create_sheet('Template_data', 2)
    WholePathList = [fpath for fpath in filenamelist]
    Wholetargetpathlist = []
    dictTargetPathlist = []

    # Get only the last three path elements
    for Pth in WholePathList:
        temp_list = Pth.split('\\')  # Use PathLib
        Wholetargetpathlist.append('\\'.join(temp_list[-3:]))
    for ppth in Wholetargetpathlist:
        path_part = '\\'.join(ppth.split('\\')[:-1])
        try:
            filename_part = file_key_mask.search(ppth).group(0)
        except AttributeError:
            print("The provided Filemask does not retrieve the given filename. "
                  "Please adjust the REGEX mask in the json file according to the image names.")
            sys.exit(1)
        dictTargetPathlist.append('\\'.join([path_part, filename_part]))
    
    # Extraction and group assignment
    subject_group = []
    area_group = []
    time_group = []
    for filpth in filenamelist:
        subjtemplist = filpth.split('\\')
        subject_group.append(subj.search(subjtemplist[-1]).group(0))

    sub_sorted = sorted(set(subject_group))

    # Tuple attaches subject number to transfer list element
    key_list = []
    for i, subno in enumerate(subject_group):
        key_list.append((dictTargetPathlist[i], subno))

    # Dictionary maps key_list with filenames
    Main_mapping_dict = {}
    for no, key in enumerate(key_list):
        Main_mapping_dict[key] = WholePathList[no]

    for row in zip(WholePathList, Wholetargetpathlist, dictTargetPathlist):
        ws2.append(row)
    wb.save(excelfile)

    ######################  SHEET 2b  #########################################################
    ws2b = wb.create_sheet('First_Subject_tempData', 3)
    sub1Pathlist = [fpath for fpath in filenamelist if sub_sorted[0] in fpath]
    targetpathlist = []
    targetdictpathlist = []

    for imagepath in sub1Pathlist:
        temp_list = imagepath.split('\\')  # Use PathLib
        filename_ele = file_key_mask.search(imagepath.split('\\')[-1]).group(0)  # Use PathLib
        path_ele = '\\'.join(temp_list[-3:-1])
        targetpathlist.append(('\\'.join(temp_list[-3:]))[:-4].replace(sub_sorted[0], "S$Sub_id$"))
        targetdictpathlist.append('\\'.join([path_ele, filename_ele.replace(sub_sorted[0], "S$Sub_id$")]))

    for row in zip(sub1Pathlist, targetpathlist, targetdictpathlist):
        ws2b.append(row)
    wb.save(excelfile)

    ######################  SHEET 3  ##########################################################
    ws3 = wb.create_sheet('Std_Template', 4)

    for target in targetpathlist:
        tempList = target.split('\\')
        area_group.append(area.search(tempList[-1]).group(0))
        time_group.append(time.search(tempList[-1]).group(0))

    area_sorted = sorted(set(area_group))
    time_sorted = sorted(set(time_group))

    area_code_list = []
    for k in area_sorted:
        area_code_list.append(dict_code[k])

    time_code_list = []
    for k in time_sorted:
        time_code_list.append(dict_code[k])

    # Writing the template to columns and rows
    for an, ac in enumerate(area_code_list, start=2):
        ws3.cell(row=an, column=1).value = ac

    for tn, tc in enumerate(time_code_list, start=2):
        ws3.cell(row=1, column=tn).value = tc

    transfer_list = []
    for sub in sub_sorted:
        for target in Wholetargetpathlist:
            if sub in target:
                target_trimmed = target[:-4]
                transfer_list.append(target_trimmed)

    transferIter = iter(transfer_list)

    for an, ac in enumerate(area_code_list, start=2):
        for tn, tc in enumerate(time_code_list, start=2):
            ws3.cell(row=an, column=tn).value = next(transferIter)
    wb.save(excelfile)

    # Make a copy of the template for user modification
    source = wb['Std_Template']
    target = wb.copy_worksheet(source)
    target.title = "Modify_Template_here"
    wb.save(excelfile)

    filenamelist_return = [os.path.split(row)[-1] for row in filenamelist]

    return area_code_list, time_code_list, area_sorted, time_sorted, sub_sorted, filenamelist, filenamelist_return, Main_mapping_dict, transfer_list
