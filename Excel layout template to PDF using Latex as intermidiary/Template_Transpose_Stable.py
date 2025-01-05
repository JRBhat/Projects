from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook
from typing import List, Iterator, Tuple

from Latex_File_Create_Stable import create_final_latex_file
from Common_Functions_Stable import create_Latex_document


def transpositioned(
    area_code_list: List[str],
    time_code_list: List[str],
    area_sorted: List[str],
    time_sorted: List[str],
    sub_sorted: List[str],
    filepaths: List[str],
    filenamelist: List[str],
    rownames: List[str],
    colnames: List[str],
    transfer_list: List[str],
    random_iter: Iterator[Tuple[str, str]],
    Imp: object,
    random: bool = True
) -> str:
    """
    Transpose of the standard template - Rows and columns are swapped in the Excel worksheet.

    This function reads an Excel file, transposes a standard template (swapping rows and columns), 
    and generates a LaTeX file based on the modified template. The function also allows for randomization 
    depending on the value of the `random` parameter.

    Args:
        area_code_list (List[str]): List of area codes.
        time_code_list (List[str]): List of time codes.
        area_sorted (List[str]): Sorted list of area names.
        time_sorted (List[str]): Sorted list of time points.
        sub_sorted (List[str]): Sorted list of subject IDs.
        filepaths (List[str]): List of file paths corresponding to the data.
        filenamelist (List[str]): List of filenames associated with the file paths.
        rownames (List[str]): List of row names for the LaTeX table.
        colnames (List[str]): List of column names for the LaTeX table.
        transfer_list (List[str]): List of transfer values to be used for the transposition.
        random_iter (Iterator[Tuple[str, str]]): Iterator of tuples containing randomization data (file paths and random codes).
        Imp (Imp): An instance of the `Internal_Imports_Stable` module containing necessary import details.
        random (bool): Whether to randomize the LaTeX document. Defaults to True.

    Returns:
        str: The path to the generated LaTeX file.
    """
    wb = load_workbook(Imp.excelfile)
    ws4 = wb['Modify_Template_here']

    nr_col = len(area_code_list)
    nr_row = len(time_code_list)

    max_height = round(1.0 / (1.25 * nr_row), 2)
    next_label_count = "".join(rownames).count("next")
    if next_label_count > 0:
        max_height = round((1.0 * next_label_count) / (1.25 * nr_row), 2)
    
    max_width = round(1.0 / (1.5 * nr_col), 2)

    # Transpose the transfer list
    transfer_list_transposed = []
    for time_code in time_sorted:
        for area_code in area_sorted:
            for transfer_value in transfer_list:
                if area_code in transfer_value and time_code in transfer_value:
                    transfer_list_transposed.append(transfer_value)

    Transpose_Iter = iter(transfer_list_transposed)  # Initialize iterator

    # Replace values in the modified Excel sheet with the transposed values
    for ttn, _ in enumerate(time_code_list, start=2):
        for aan, _ in enumerate(area_code_list, start=2):
            ws4.cell(row=ttn, column=aan).value = next(Transpose_Iter)

    wb.save(Imp.excelfile)

    # Create LaTeX document based on the randomization flag
    if random:
        lax_document, new_page_alert = create_Latex_document(
            colnames, rownames, sub_sorted, max_height, max_width,
            time_sorted, area_sorted, filepaths, filenamelist, 
            Type="Standard", RandomList=random_iter
        )
    else:
        lax_document, new_page_alert = create_Latex_document(
            colnames, rownames, sub_sorted, max_height, max_width,
            time_sorted, area_sorted, filepaths, filenamelist, 
            Type="Standard", RandomList=random_iter
        )

    # Generate final LaTeX file
    transpose_tex_file = create_final_latex_file(
        Imp.studynumber, Imp.header, Imp.pagestyle, 
        Imp.hypersetup, lax_document, new_page_alert, 
        colnames, Imp.Test_type, draft=Imp.draft_flag
    )

    return transpose_tex_file
