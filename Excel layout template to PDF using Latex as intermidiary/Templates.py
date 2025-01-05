from Latex_File_Create_Stable import create_final_latex_file
from Common_Functions_Stable import create_Latex_document

from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook
from typing import List, Dict


class Standard:
    """
    A class representing a standard template for generating LaTeX files.
    """

    def __init__(self, area_code_list: List[str], time_code_list: List[str], rownames: List[str]) -> None:
        """
        Initialize the Standard class.

        :param area_code_list: A list of area codes.
        :param time_code_list: A list of time codes.
        :param rownames: A list of row names.
        """
        self.max_height, self.max_width = self.get_max_height_and_width(area_code_list, time_code_list, rownames)

    def get_max_height_and_width(self, area_code_list: List[str], time_code_list: List[str], rownames: List[str]) -> tuple:
        """
        Calculate the maximum height and width for the LaTeX document.

        :param area_code_list: A list of area codes.
        :param time_code_list: A list of time codes.
        :param rownames: A list of row names.
        :return: A tuple containing the maximum height and width.
        """
        nr_row = len(area_code_list)
        nr_col = len(time_code_list)

        max_height = round(1.0 / (1.25 * nr_row), 2)
        next_label_count = "".join(rownames).count("next")
        if next_label_count > 0:
            max_height = round((1.0 * next_label_count) / (1.25 * nr_row), 2)
        max_width = round(1.0 / (1.5 * nr_col), 2)
        return max_height, max_width

    def get_final_tex_file(self, 
                           sub_sorted: List[str], area_sorted: List[str], time_sorted: List[str],
                           rownames: List[str], colnames: List[str], 
                           filepaths: List[str], filenamelist: List[str], 
                           random_iter: List[str], Imp: object, 
                           Type: str = "Standard") -> str:
        """
        Generate the final LaTeX file.

        :param sub_sorted: A sorted list of subjects.
        :param area_sorted: A sorted list of areas.
        :param time_sorted: A sorted list of times.
        :param rownames: A list of row names.
        :param colnames: A list of column names.
        :param filepaths: A list of file paths.
        :param filenamelist: A list of file names.
        :param random_iter: A list of random iterations.
        :param Imp: An object containing the import settings (e.g., study number, header, etc.).
        :param Type: The type of template ("Standard" by default).
        :return: The generated LaTeX file as a string.
        """
        lax_document, new_page_alert = create_Latex_document(colnames,  rownames,
                                                              sub_sorted, 
                                                              self.max_height, self.max_width, 
                                                              area_sorted, time_sorted, 
                                                              filepaths, filenamelist, 
                                                              Type, 
                                                              RandomList=random_iter)

        tex_file = create_final_latex_file(Imp.studynumber, Imp.header, Imp.pagestyle, Imp.hypersetup, 
                                           lax_document, new_page_alert, 
                                           colnames, Imp.Test_type, 
                                           draft=Imp.draft_flag)

        return tex_file


class Transpose(Standard):
    """
    A subclass of Standard that implements transposing the area and time codes.
    """

    def __init__(self, area_code_list: List[str], time_code_list: List[str], 
                 area_sorted: List[str], time_sorted: List[str], 
                 rownames: List[str], transfer_list: List[str], 
                 Imp: object, random: bool = True) -> None:
        """
        Initialize the Transpose class, swapping area and time codes and setting up the template.

        :param area_code_list: A list of area codes.
        :param time_code_list: A list of time codes.
        :param area_sorted: A sorted list of areas.
        :param time_sorted: A sorted list of times.
        :param rownames: A list of row names.
        :param transfer_list: A list of transfer values.
        :param Imp: An object containing the import settings.
        :param random: A flag indicating whether to apply randomization (default is True).
        """
        area_code_list_swapped = time_code_list
        time_code_list_swapped = area_code_list

        self.max_height, self.max_width = super().get_max_height_and_width(area_code_list_swapped, time_code_list_swapped, rownames)
        self.transpose_excel_template(area_code_list, time_code_list, area_sorted, time_sorted, transfer_list, Imp)

    def transpose_excel_template(self, area_code_list: List[str], time_code_list: List[str], 
                                 area_sorted: List[str], time_sorted: List[str], 
                                 transfer_list: List[str], Imp: object) -> None:
        """
        Modify an Excel template based on the transposed area and time codes.

        :param area_code_list: A list of area codes.
        :param time_code_list: A list of time codes.
        :param area_sorted: A sorted list of areas.
        :param time_sorted: A sorted list of times.
        :param transfer_list: A list of transfer values.
        :param Imp: An object containing the import settings.
        """
        wb = load_workbook(Imp.excelfile)
        ws4 = wb['Modify_Template_here']

        transfer_list_transposed = []
        for tcode in time_sorted:
            for acode in area_sorted:
                for tvalue in transfer_list:
                    if acode in tvalue and tcode in tvalue:
                        transfer_list_transposed.append(tvalue)

        Transpose_Iter = iter(transfer_list_transposed)  # Initialize iterator

        for ttn, _ in enumerate(time_code_list, start=2):  
            for aan, _ in enumerate(area_code_list, start=2):  
                ws4.cell(row=ttn, column=aan).value = next(Transpose_Iter)

        wb.save(Imp.excelfile)

    def get_final_tex_file(self, 
                           sub_sorted: List[str], area_sorted: List[str], time_sorted: List[str],
                           rownames: List[str], colnames: List[str], 
                           filepaths: List[str], filenamelist: List[str], 
                           random_iter: List[str], Imp: object, 
                           Type: str = "Standard") -> str:
        """
        Generate the final LaTeX file with transposed area and time codes.

        :param sub_sorted: A sorted list of subjects.
        :param area_sorted: A sorted list of areas.
        :param time_sorted: A sorted list of times.
        :param rownames: A list of row names.
        :param colnames: A list of column names.
        :param filepaths: A list of file paths.
        :param filenamelist: A list of file names.
        :param random_iter: A list of random iterations.
        :param Imp: An object containing the import settings.
        :param Type: The type of template ("Standard" by default).
        :return: The generated LaTeX file as a string.
        """
        lax_document, new_page_alert = create_Latex_document(colnames, rownames,
                                                              sub_sorted, 
                                                              self.max_height, self.max_width, 
                                                              time_sorted, area_sorted, 
                                                              filepaths, filenamelist, 
                                                              Type, 
                                                              RandomList=random_iter)

        tex_file = create_final_latex_file(Imp.studynumber, Imp.header, Imp.pagestyle, Imp.hypersetup, 
                                           lax_document, new_page_alert, 
                                           colnames, Imp.Test_type, 
                                           draft=Imp.draft_flag)

        return tex_file


class Custom(Standard):
    """
    A subclass of Standard that implements a custom template.
    """

    def __init__(self, rownames: List[str], colnames: List[str]) -> None:
        """
        Initialize the Custom class with the row and column names for a custom template.

        :param rownames: A list of row names.
        :param colnames: A list of column names.
        """
        self.max_height, self.max_width = self.get_max_height_and_width(rownames, colnames)

    def get_max_height_and_width(self, rownames: List[str], colnames: List[str]) -> tuple:
        """
        Calculate the maximum height and width for the custom LaTeX document.

        :param rownames: A list of row names.
        :param colnames: A list of column names.
        :return: A tuple containing the maximum height and width.
        """
        next_label_count = "".join(rownames).count("next")
        nr_col = len(colnames)
        nr_row = len(rownames)

        max_height = round(1.0 / (1.15 * nr_row), 2)
        if next_label_count > 0:
            max_height = round((1.0 * next_label_count) / (1.25 * nr_row), 2)

        max_width = round(1.0 / (1.2 * nr_col), 2)
        return max_height, max_width

    def get_final_tex_file(self, sub_sorted: List[str], area_sorted: List[str], time_sorted: List[str],
                           rownames: List[str], colnames: List[str], 
                           main_mapping_dict: Dict[str, str], 
                           random_iter: List[str], Imp: object) -> str:
        """
        Generate the final LaTeX file for a custom template.

        :param sub_sorted: A sorted list of subjects.
        :param area_sorted: A sorted list of areas.
        :param time_sorted: A sorted list of times.
        :param rownames: A list of row names.
        :param colnames: A list of column names.
        :param main_mapping_dict: A dictionary mapping values in the LaTeX document.
        :param random_iter: A list of random iterations.
        :param Imp: An object containing the import settings.
        :return: The generated LaTeX file as a string.
        """
        ws4 = load_workbook(Imp.excelfile)['Modify_Template_here']
        lax_document, new_page_alert = create_Latex_document(colnames, rownames,
                                                              sub_sorted, 
                                                              self.max_height, self.max_width, 
                                                              area_sorted, time_sorted, 
                                                              ws4, main_mapping_dict, 
                                                              Type="Custom", 
                                                              RandomList=random_iter)

        tex_file = create_final_latex_file(Imp.studynumber, Imp.header, Imp.pagestyle, Imp.hypersetup, 
                                           lax_document, new_page_alert, 
                                           colnames, Imp.Test_type, 
                                           draft=Imp.draft_flag)

        return tex_file
